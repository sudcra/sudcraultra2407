from eval_insert import txt_a_df , inserta_archivo, ejecutasqlarch, ejecutasql, hace_conexion , cierra_conexion
from rutinas2.admin_df_prueba2 import convertir_a_json
from rutinas2.admin_4 import generar_html_docente, generar_html_estudiantes
from agrega_registros import agregar_registros
from datetime import datetime, timezone
import pandas as pd
import psycopg2
from psycopg2 import sql
from crea_pptx import creappt
import xlsxwriter
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import quote_sheetname
import xlwings as xw
from openpyxl.workbook.protection import WorkbookProtection
import os


    

def copiasPlanillas(archivox, cod_asig, carpeta, sufijo):
    raiz, extension = os.path.splitext(archivox)
    path_base='C:/sudcraultra/Consultas/'

    archivo = open(path_base + 'listado_planillas.sql', "r")
    sql= archivo.read()
    sql = sql.replace("[cod_asig]", cod_asig)
    df=ejecutasql(sql)
    i=1
    total_registros = len(df)
    with tqdm(total=total_registros, desc="Progreso planillas") as pbar:
        for row in df.itertuples():
            archivo = open(path_base + 'listado_alumnos_planillas.sql', "r")
            sql= archivo.read()
            sql = sql.replace("[id_seccion]", str(row.id_seccion))
            dfa=ejecutasql(sql)
            ruta_completa = f"{carpeta}/{row.cod_programa}/{row.cod_sede}/{sufijo}/{cod_asig}"
  
            try:
                os.makedirs(ruta_completa)
            except Exception as e:
                m=0
            else:
                m=0
            finally:
                m=0
            new_planilla = f"{carpeta}/{row.cod_programa}/{row.cod_sede}/{sufijo}/{cod_asig}/{row.cod_sede}_{row.seccion}_{sufijo}{extension}"
            dfAplanilla(dfa,archivox,new_planilla)
            pbar.update(1)



def dfAplanilla(df, excel_file ,new_excel_file):
    # Cargar el archivo Excel existente
    #wb = load_workbook(filename=excel_file, read_only=False)
    wb = load_workbook(filename=excel_file, read_only=False, keep_vba=True)
    ws = wb.active
   
    # Obtener la posición de la celda (10,2)
    start_row = 10
    start_col = 2

    # Pega el DataFrame a partir de la celda (10,2)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            ws.cell(row=r_idx, column=c_idx, value=value)

            # Elimina las últimas filas

    for fila in range (10 + len(df), 70):  
         ws.row_dimensions[fila].hidden = True


    #while ws.max_row > 10 + len(df)-1:
        #ws.delete_rows(ws.max_row)
        #   ws.row_dimensions[ws.max_row].hidden = True
        


    # Proteger la hoja con la contraseña "arcdus"
    ws.protection.sheet = True
    ws.protection.set_password('arcdus')



    wsh = wb["Hoja1"]
    wsh.sheet_state = 'hidden'
    
    
    
    wb.security = WorkbookProtection(workbookPassword = 'arcdus', lockStructure = True)
    

    # Guarda con un nombre diferente
    
    wb.save(new_excel_file)
    wb.close()
    
if __name__ == "__main__":

   
# MATEMÁTICA
    if 1==2 :
        cod_asig='MAT3110'
        prueba=4
        ruta='C:/sudcraultra_access/SISTEMA/'
        archivo=f'{cod_asig}_{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/Fundacion Instituto Profesional Duoc UC/Docentes Programa Matemática DUOC UC - PLANILLAS/MAT20241/PLANILLAS'
        sufijo='ET'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo)   
    
 # EMPRENDIMIENTO   
    if 1==2 :
        cod_asig='EMP1101'
        prueba=4
        ruta='C:/sudcraultra_access/SISTEMA/'
        archivo=f'{cod_asig}_{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/EMP20241/PLANILLAS'
        sufijo='ET'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo)
 # ETICA 
    if 1==2 :
        cod_asig='EAT6845'
        prueba=2
        ruta='C:/sudcraultra_access/SISTEMA/'
        archivo=f'{cod_asig}_{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/EYFC20241/PLANILLAS'
        sufijo='E2'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo)


# LENGUAJE 
    if 1==2 :
        cod_asig='PLC1101'
        prueba=4
        ruta='C:/sudcraultra_access/SISTEMA/'
        archivo=f'{cod_asig}_{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/LEN20241'
        sufijo='ET'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo)

# INGLES 
    if 1==2 :
        cod_asig='INI7111'
        prueba=1
        ruta='C:/sudcraultra_access/SISTEMA/'
        archivo=f'{cod_asig}_{prueba}.xlsm'
        rutaArchivo= ruta + archivo
        carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/ING20241'
        sufijo='ET_ORAL'
        copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo)


    if 1==2:
        path_base='C:/sudcraultra/Consultas/'
        # Consula SQL tabla informes_secciones_pendientes
        sql="select distinct e.cod_asig from eval e join asignaturas asig on asig.cod_asig = e.cod_asig where cod_programa = 'pemp' and num_prueba = 4"
        df=ejecutasql(sql)
        # Consulta SQL tabla 

        i=1
        total_registros = len(df)
        with tqdm(total=total_registros, desc="Progreso informes secciones") as pbar:
            for row in df.itertuples():

                cod_asig=row.cod_asig
                prueba=4
                print(cod_asig)
                ruta='C:/sudcraultra_access/SISTEMA/'
                archivo=f'{cod_asig}_{prueba}.xlsm'
                rutaArchivo= ruta + archivo
                carp = 'C:/Users/lgutierrez/OneDrive - Fundacion Instituto Profesional Duoc UC/SUDCRA/PLANILLAS/EMP20241/PLANILLAS'
                sufijo='ET'
                copiasPlanillas(rutaArchivo,cod_asig,carp,sufijo)
                pbar.update(1)